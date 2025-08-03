import pandas as pd
import json
from pathlib import Path
from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
import logging

class ExcelHandler:
    def __init__(self):
        from ..config import Config
        self.config = Config()
        self.logger = logging.getLogger(__name__)
    
    def save_individual_invoice(self, invoice_data: Dict[str, Any], invoice_number: str) -> str:
        """Save individual invoice data to Excel file"""
        try:
            # Clean invoice number for filename
            clean_invoice_number = self._clean_filename(invoice_number)
            filename = f"invoice_{clean_invoice_number}.xlsx"
            filepath = self.config.INDIVIDUAL_INVOICES_DIR / filename
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Sheet 1: Invoice Summary
                self._write_invoice_summary_sheet(writer, invoice_data)
                
                # Sheet 2: Line Items Details
                self._write_line_items_sheet(writer, invoice_data)
                
                # Sheet 3: Raw Data (for debugging)
                self._write_raw_data_sheet(writer, invoice_data)
            
            self.logger.info(f"Saved individual invoice Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Error saving individual invoice Excel: {e}")
            raise
    
    def _write_invoice_summary_sheet(self, writer, invoice_data: Dict[str, Any]):
        """Write invoice summary to Sheet1"""
        summary_data = []
        
        # Invoice Information
        summary_data.append(['INVOICE INFORMATION', '', ''])
        invoice_info = invoice_data.get('invoice_info', {})
        for key, value in invoice_info.items():
            summary_data.append(['Invoice Info', key.replace('_', ' ').title(), value])
        
        summary_data.append(['', '', ''])  # Empty row
        
        # Vendor Information
        summary_data.append(['VENDOR INFORMATION', '', ''])
        vendor = invoice_data.get('vendor', {})
        for key, value in vendor.items():
            summary_data.append(['Vendor', key.replace('_', ' ').title(), value])
        
        summary_data.append(['', '', ''])  # Empty row
        
        # Customer Information
        summary_data.append(['CUSTOMER INFORMATION', '', ''])
        customer = invoice_data.get('customer', {})
        for key, value in customer.items():
            summary_data.append(['Customer', key.replace('_', ' ').title(), value])
        
        summary_data.append(['', '', ''])  # Empty row
        
        # Totals
        summary_data.append(['TOTALS', '', ''])
        totals = invoice_data.get('totals', {})
        for key, value in totals.items():
            summary_data.append(['Totals', key.replace('_', ' ').title(), value])
        
        summary_data.append(['', '', ''])  # Empty row
        
        # Payment Information
        summary_data.append(['PAYMENT INFORMATION', '', ''])
        payment_info = invoice_data.get('payment_info', {})
        for key, value in payment_info.items():
            summary_data.append(['Payment Info', key.replace('_', ' ').title(), value])
        
        # Create DataFrame and save
        df = pd.DataFrame(summary_data, columns=['Section', 'Field', 'Value'])
        df.to_excel(writer, sheet_name='Invoice Summary', index=False)
        
        # Format the sheet
        self._format_summary_sheet(writer.sheets['Invoice Summary'])
    
    def _write_line_items_sheet(self, writer, invoice_data: Dict[str, Any]):
        """Write line items to separate sheet"""
        line_items = invoice_data.get('line_items', [])
        
        if not line_items:
            # Create empty sheet with headers
            df_items = pd.DataFrame(columns=[
                'Item Name', 'Item Code', 'HSN/SAC Code', 'Quantity',
                'Taxable Amount', 'Discount Amount', 'CGST Rate', 'CGST Amount',
                'SGST Rate', 'SGST Amount', 'IGST Rate', 'IGST Amount',
                'CESS Rate', 'CESS Amount', 'Total Amount'
            ])
        else:
            # Convert line items to DataFrame
            formatted_items = []
            for item in line_items:
                formatted_item = {
                    'Item Name': item.get('item_name', ''),
                    'Item Code': item.get('item_code', ''),
                    'HSN/SAC Code': item.get('hsn_sac_code', ''),
                    'Quantity': item.get('quantity', 0),
                    'Taxable Amount': item.get('taxable_amount', 0),
                    'Discount Amount': item.get('discount_amount', 0),
                    'CGST Rate': item.get('cgst_rate', 0),
                    'CGST Amount': item.get('cgst_amount', 0),
                    'SGST Rate': item.get('sgst_rate', 0),
                    'SGST Amount': item.get('sgst_amount', 0),
                    'IGST Rate': item.get('igst_rate', 0),
                    'IGST Amount': item.get('igst_amount', 0),
                    'CESS Rate': item.get('cess_rate', 0),
                    'CESS Amount': item.get('cess_amount', 0),
                    'Total Amount': item.get('total_amount', 0)
                }
                formatted_items.append(formatted_item)
            
            df_items = pd.DataFrame(formatted_items)
        
        df_items.to_excel(writer, sheet_name='Line Items', index=False)
        
        # Format the sheet
        self._format_line_items_sheet(writer.sheets['Line Items'])
    
    def _write_raw_data_sheet(self, writer, invoice_data: Dict[str, Any]):
        """Write raw extracted data for debugging"""
        # Convert to a readable format
        raw_data = []
        
        def flatten_dict(d, parent_key='', sep='_'):
            items = []
            for k, v in d.items():
                new_key = f"{parent_key}{sep}{k}" if parent_key else k
                if isinstance(v, dict):
                    items.extend(flatten_dict(v, new_key, sep=sep).items())
                elif isinstance(v, list):
                    for i, item in enumerate(v):
                        if isinstance(item, dict):
                            items.extend(flatten_dict(item, f"{new_key}_{i}", sep=sep).items())
                        else:
                            items.append((f"{new_key}_{i}", item))
                else:
                    items.append((new_key, v))
            return dict(items)
        
        flattened = flatten_dict(invoice_data)
        for key, value in flattened.items():
            raw_data.append([key, str(value)])
        
        df_raw = pd.DataFrame(raw_data, columns=['Field', 'Value'])
        df_raw.to_excel(writer, sheet_name='Raw Data', index=False)
    
    def update_master_excel(self, mapped_data: Dict[str, Any]) -> str:
        """Update or create master Excel file with mapped data"""
        try:
            filepath = self.config.MASTER_DATA_DIR / self.config.MASTER_EXCEL_FILE
            
            # Convert mapped data to DataFrame row
            df_new_row = pd.DataFrame([mapped_data])
            
            if filepath.exists():
                # Read existing data
                try:
                    df_existing = pd.read_excel(filepath)
                    # Ensure columns match
                    for col in df_new_row.columns:
                        if col not in df_existing.columns:
                            df_existing[col] = ""
                    for col in df_existing.columns:
                        if col not in df_new_row.columns:
                            df_new_row[col] = ""
                    
                    # Reorder columns to match
                    df_new_row = df_new_row[df_existing.columns]
                    df_combined = pd.concat([df_existing, df_new_row], ignore_index=True)
                except Exception as e:
                    self.logger.warning(f"Error reading existing master file: {e}. Creating new file.")
                    df_combined = df_new_row
            else:
                df_combined = df_new_row
            
            # Save updated data
            df_combined.to_excel(filepath, index=False)
            
            # Format the master Excel
            self._format_master_excel(filepath)
            
            self.logger.info(f"Updated master Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            self.logger.error(f"Error updating master Excel: {e}")
            raise
    
    def _format_summary_sheet(self, worksheet):
        """Format the invoice summary sheet"""
        try:
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Section header formatting
            section_font = Font(bold=True, size=12)
            section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.row == 1:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal="center")
                    elif cell.value and isinstance(cell.value, str) and cell.value.isupper():
                        # Section headers (all caps)
                        cell.font = section_font
                        cell.fill = section_fill
            
            # Adjust column widths
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 25
            worksheet.column_dimensions['C'].width = 30
            
        except Exception as e:
            self.logger.warning(f"Error formatting summary sheet: {e}")
    
    def _format_line_items_sheet(self, worksheet):
        """Format the line items sheet"""
        try:
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            
            for cell in worksheet[1]:  # First row (header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)  # Cap at 30
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.logger.warning(f"Error formatting line items sheet: {e}")
    
    def _format_master_excel(self, filepath: Path):
        """Format the master Excel file"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="C55A5A", end_color="C55A5A", fill_type="solid")
            
            for cell in ws[1]:  # First row (header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)  # Cap at 25
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            
        except Exception as e:
            self.logger.warning(f"Error formatting master Excel: {e}")
    
    def _clean_filename(self, filename: str) -> str:
        """Clean filename to remove invalid characters"""
        import re
        return re.sub(r'[<>:"/\\|?*]', '_', str(filename))
    
    def get_master_excel_stats(self) -> Dict[str, Any]:
        """Get statistics about the master Excel file"""
        try:
            filepath = self.config.MASTER_DATA_DIR / self.config.MASTER_EXCEL_FILE
            
            if not filepath.exists():
                return {
                    "exists": False,
                    "total_invoices": 0,
                    "file_size": 0
                }
            
            df = pd.read_excel(filepath)
            
            return {
                "exists": True,
                "total_invoices": len(df),
                "file_size": filepath.stat().st_size,
                "file_size_formatted": self._format_file_size(filepath.stat().st_size),
                "columns": list(df.columns),
                "last_modified": filepath.stat().st_mtime
            }
            
        except Exception as e:
            self.logger.error(f"Error getting master Excel stats: {e}")
            return {"exists": False, "error": str(e)}
    
    @staticmethod
    def _format_file_size(size_bytes: int) -> str:
        """Format file size in human readable format"""
        if size_bytes == 0:
            return "0B"
        
        size_names = ["B", "KB", "MB", "GB"]
        import math
        i = int(math.floor(math.log(size_bytes, 1024)))
        p = math.pow(1024, i)
        s = round(size_bytes / p, 2)
        return f"{s} {size_names[i]}"
